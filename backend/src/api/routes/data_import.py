import re
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import text

from api.auth import decode_access_token
from db.database import get_session

router = APIRouter(prefix="/api/data-import", tags=["数据导入"])


def get_current_user_from_header(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="未提供认证信息")

    token = authorization.replace("Bearer ", "")
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Token已过期或无效")

    return payload


@router.get("/sprints/{project_id}")
async def list_closed_sprints(
    project_id: int,
    current_user: dict = Depends(get_current_user_from_header)
):
    session = get_session()

    try:
        query = text("SELECT project_id FROM project_configs WHERE id = :project_id")
        result = session.execute(query, {"project_id": project_id})
        row = result.fetchone()

        if not row:
            return []

        jira_project_id = row[0]

        sprints_query = text("""
            SELECT sprint_id, sprint_name
            FROM rdm_sprint
            WHERE project_id = :jira_project_id
            AND state = 'closed'
            GROUP BY sprint_id, sprint_name
            ORDER BY sprint_id DESC
        """)
        sprints_result = session.execute(sprints_query, {"jira_project_id": jira_project_id})

        sprints = []
        for sprint_row in sprints_result:
            sprints.append({
                "sprint_id": sprint_row[0],
                "sprint_name": sprint_row[1],
            })
        return sprints
    finally:
        session.close()


class DocBugItem(BaseModel):
    key: str | None = None
    name: str | None = None
    priority: str | None = None
    reason: str | None = None
    resolve_method: str | None = None
    maker: str | None = None
    propose: str | None = None
    propose_time: datetime | None = None
    resolve_time: datetime | None = None
    status: str | None = None
    type: str | None = None
    original_type: str | None = None
    sprint_id: str | None = None
    project_id: str | None = None


@router.get("/doc-bugs")
async def list_doc_bugs(
    project_id: str = Query(..., description="JIRA项目ID"),
    sprint_id: str = Query(..., description="Sprint ID"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user_from_header)
):
    session = get_session()

    try:
        count_query = text("""
            SELECT COUNT(*) FROM rdm_doc_bug
            WHERE project_id = :project_id AND sprint_id = :sprint_id
        """)
        count_result = session.execute(count_query, {
            "project_id": project_id,
            "sprint_id": sprint_id,
        })
        total = count_result.fetchone()[0]

        data_query = text("""
            SELECT `key`, `name`, priority, reason, resolve_method,
                   maker, propose, propose_time, resolve_time,
                   status, `type`, original_type, sprint_id, project_id
            FROM rdm_doc_bug
            WHERE project_id = :project_id AND sprint_id = :sprint_id
            ORDER BY `key`
            LIMIT :limit OFFSET :offset
        """)
        data_result = session.execute(data_query, {
            "project_id": project_id,
            "sprint_id": sprint_id,
            "limit": page_size,
            "offset": (page - 1) * page_size,
        })

        items = []
        for row in data_result:
            items.append({
                "key": row[0],
                "name": row[1],
                "priority": row[2],
                "reason": row[3],
                "resolve_method": row[4],
                "maker": row[5],
                "propose": row[6],
                "propose_time": row[7].isoformat() if row[7] else None,
                "resolve_time": row[8].isoformat() if row[8] else None,
                "status": row[9],
                "type": row[10],
                "original_type": row[11],
                "sprint_id": str(row[12]) if row[12] else None,
                "project_id": str(row[13]) if row[13] else None,
            })

        return {"total": total, "page": page, "page_size": page_size, "items": items}
    finally:
        session.close()


class ClearDocBugsRequest(BaseModel):
    project_id: str
    sprint_id: str


@router.delete("/doc-bugs")
async def clear_doc_bugs(
    request: ClearDocBugsRequest,
    current_user: dict = Depends(get_current_user_from_header)
):
    session = get_session()

    try:
        delete_query = text("""
            DELETE FROM rdm_doc_bug
            WHERE project_id = :project_id AND sprint_id = :sprint_id
        """)
        result = session.execute(delete_query, {
            "project_id": request.project_id,
            "sprint_id": request.sprint_id,
        })
        session.commit()
        return {"deleted": result.rowcount}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"清除数据失败: {str(e)}")
    finally:
        session.close()


PRIORITY_MAP = {"极高": "致命", "高": "严重", "中": "一般", "低": "优化"}

VALID_STATUSES = {"不处理", "处理中", "待测试", "继续观察", "未开始", "验证通过", "已解决", "转任务", "转需求"}

TYPE_MAP = {
    "代码逻辑问题": "代码实现",
    "漏做": "代码实现",
    "需求问题": "业务需求",
    "环境问题": "环境配置",
    "第三方影响": "其他",
    "沟通问题": "其他",
}

VALID_ORIGINAL_TYPES = {"需求问题", "漏做", "代码逻辑问题", "环境问题", "第三方影响", "沟通问题"}


def parse_maker(raw: str | None) -> str:
    if not raw:
        return ""
    cleaned = raw.replace("@", ",").replace("\n", ",").replace("，", ",").replace("\r", ",")
    parts = [p.strip() for p in cleaned.split(",")]
    parts = [p for p in parts if p]
    return parts[-1] if parts else ""


def parse_datetime(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d",
            "%Y-%m-%dT%H:%M:%S",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def find_column_index(headers: list[str], candidates: list[str]) -> int | None:
    for candidate in candidates:
        for i, h in enumerate(headers):
            if h and h.strip() == candidate:
                return i
    return None


def find_column_index_contains(headers: list[str], keyword: str) -> int | None:
    for i, h in enumerate(headers):
        if h and keyword in h:
            return i
    return None


@router.post("/doc-bugs/upload")
async def upload_doc_bugs(
    project_id: str = Query(..., description="JIRA项目ID"),
    sprint_id: str = Query(..., description="Sprint ID"),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user_from_header)
):
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Excel 文件至少需要包含表头和一行数据")

        headers = [str(h).strip() if h else "" for h in rows[0]]

        key_col = find_column_index(headers, ["问题编号"])
        func_col = find_column_index(headers, ["功能点"])
        desc_col = find_column_index(headers, ["问题详述"])
        priority_col = find_column_index(headers, ["优先级"])
        reason_col = find_column_index(headers, ["原因分析"])
        opinion_col = find_column_index(headers, ["处理意见"])
        method_col = find_column_index(headers, ["处理方式"])
        maker_col = find_column_index(headers, ["处理人"])
        propose_col = find_column_index(headers, ["提出人"])
        propose_time_col = find_column_index(headers, ["提出时间"])
        resolve_time_col = find_column_index(headers, ["完成时间"])
        status_col = find_column_index(headers, ["处理状态"])
        type_col = find_column_index(headers, ["问题类型"])

        errors = []
        valid_rows = []

        for row_idx, row in enumerate(rows[1:], start=2):
            row_errors = []

            key_val = str(row[key_col]).strip() if key_col is not None and row[key_col] is not None else ""
            if not key_val:
                row_errors.append("[问题编号]为空")
                errors.append({"row": row_idx, "errors": row_errors})
                continue

            desc_val = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] is not None else ""
            if not desc_val:
                row_errors.append("[问题详述]为空")

            func_val = str(row[func_col]).strip() if func_col is not None and row[func_col] is not None else ""
            if func_val:
                name_val = f"{func_val}>{desc_val}"
            else:
                name_val = desc_val

            priority_val = str(row[priority_col]).strip() if priority_col is not None and row[priority_col] is not None else ""
            if priority_val not in PRIORITY_MAP:
                if priority_val:
                    row_errors.append(f"[优先级] '{priority_val}' 不在 极高/高/中/低 范围内")
                priority_mapped = ""
            else:
                priority_mapped = PRIORITY_MAP[priority_val]

            reason_val = str(row[reason_col]).strip() if reason_col is not None and row[reason_col] is not None else ""
            opinion_val = str(row[opinion_col]).strip() if opinion_col is not None and row[opinion_col] is not None else ""
            if reason_val:
                reason_final = f"{reason_val}>{opinion_val}" if opinion_val else reason_val
            else:
                reason_final = opinion_val

            method_val = str(row[method_col]).strip() if method_col is not None and row[method_col] is not None else ""
            maker_raw = str(row[maker_col]).strip() if maker_col is not None and row[maker_col] is not None else ""
            maker_val = parse_maker(maker_raw)

            propose_val = str(row[propose_col]).strip() if propose_col is not None and row[propose_col] is not None else ""
            propose_time_val = parse_datetime(row[propose_time_col]) if propose_time_col is not None else None
            resolve_time_val = parse_datetime(row[resolve_time_col]) if resolve_time_col is not None else None

            status_raw = str(row[status_col]).strip() if status_col is not None and row[status_col] is not None else ""
            if status_raw and status_raw not in VALID_STATUSES:
                row_errors.append(f"[处理状态] '{status_raw}' 不在有效范围内, 有效值: {', '.join(VALID_STATUSES)}")

            type_raw = str(row[type_col]).strip() if type_col is not None and row[type_col] is not None else ""
            if type_raw and type_raw not in VALID_ORIGINAL_TYPES:
                row_errors.append(f"[问题类型] '{type_raw}' 不在有效范围内, 有效值: {', '.join(VALID_ORIGINAL_TYPES)}")
            type_mapped = TYPE_MAP.get(type_raw, "")

            if row_errors:
                errors.append({"row": row_idx, "errors": row_errors})
                continue

            valid_rows.append({
                "key": key_val,
                "name": name_val,
                "priority": priority_mapped,
                "reason": reason_final,
                "resolve_method": method_val,
                "maker": maker_val,
                "propose": propose_val,
                "propose_time": propose_time_val,
                "resolve_time": resolve_time_val,
                "status": status_raw,
                "type": type_mapped,
                "original_type": type_raw,
                "sprint_id": sprint_id,
                "project_id": project_id,
            })

        if errors:
            error_messages = []
            for err in errors:
                error_messages.append(f"第{err['row']}行: {'; '.join(err['errors'])}")
            return {
                "success": False,
                "errors": error_messages,
                "valid_count": len(valid_rows),
                "total_count": len(rows) - 1,
            }

        if not valid_rows:
            raise HTTPException(status_code=400, detail="没有有效数据可导入")

        session = get_session()
        try:
            insert_sql = text("""
                INSERT INTO rdm_doc_bug
                (`key`, `name`, priority, reason, resolve_method,
                 maker, propose, propose_time, resolve_time,
                 status, `type`, original_type, sprint_id, project_id)
                VALUES
                (:key, :name, :priority, :reason, :resolve_method,
                 :maker, :propose, :propose_time, :resolve_time,
                 :status, :type, :original_type, :sprint_id, :project_id)
            """)

            for row_data in valid_rows:
                session.execute(insert_sql, row_data)

            session.commit()

            return {
                "success": True,
                "imported": len(valid_rows),
                "total_count": len(rows) - 1,
            }
        except Exception as e:
            session.rollback()
            raise HTTPException(status_code=500, detail=f"导入数据失败: {str(e)}")
        finally:
            session.close()

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理文件失败: {str(e)}")