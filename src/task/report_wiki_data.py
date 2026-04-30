import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # 获取绝对路径,向上定位到根目录,临时添加到模块搜索路径

import requests
from base64 import b64encode
from bs4 import BeautifulSoup
import pandas as pd
from io import BytesIO
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from db.dboperator import DbOperator


# WIKI信息
CONFLUENCE_URL = 'http://wiki.zvos.zoomlion.com'
USERNAME = '00773908'
PASSWORD = 'Nx.0918@ZLZK12'

# 页面ID和项目名称
PAGE_IDS = {"67334611": "三维生产管控系统", "79220978": "无人机巡检系统", "104896273": "数据工具链平台"}

# 创建基本认证头部
auth = b64encode(f"{USERNAME}:{PASSWORD}".encode()).decode()
headers = {
    'Authorization': f'Basic {auth}',
    'Content-Type': 'application/json',
}


def parse_detail_from_xlsx(url):
    """解析评审详情Excel文件"""
    # 发送GET请求下载文件
    response = requests.get(url, auth=HTTPBasicAuth(USERNAME, PASSWORD), verify=False)  # 注意：verify=False是为了忽略SSL验证警告
    if response.status_code == 200:
        try:
            # 忽略 UserWarning
            import warnings
            warnings.filterwarnings("ignore", category=UserWarning, message="Workbook contains no default style")
            # 直接使用BytesIO处理二进制数据
            # 使用openpyxl读取Excel文件内容
            wb = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
            sheet = wb.active

            code_format = ""
            review_record_details = []
            for row in sheet.iter_rows(values_only=True):
                if sum(1 for cell in row if cell is not None) >= 5:
                    # 表头(去掉空单元格)
                    row_values_for_head = '&|&'.join(str(cell) for cell in row if cell is not None)
                    # 表体(保留空单元格)
                    row_values_for_detail = '&|&'.join(str(cell) if cell is not None else '' for cell in row )
                    if '项目信息&|&' in row_values_for_head:
                        code_format = row_values_for_head
                    else:
                        review_record_details.append(row_values_for_detail)
                else:
                    continue

            backend_code_format1 = "ID&|&项目信息&|&Git仓库&|&Git分支&|&文件路径&|&代码行号&|&代码片段&|&意见类型&|&检视人员&|&" \
                                   "检视时间&|&检视意见&|&指定确认人员&|&实际确认人员&|&确认结果&|&确认说明&|&确认时间"

            backend_code_format2 = "ID&|&项目信息&|&文件路径&|&代码行号&|&代码片段&|&意见类型&|&检视人员&|&检视时间&|&检视意见&|&" \
                                   "指定确认人员&|&实际确认人员&|&确认结果"

            frontend_code_format = "项目信息&|&文件路径&|&代码行号&|&代码片段&|&意见类型&|&检视人员&|&检视时间&|&检视意见&|&" \
                                   "实际确认人员&|&确认结果&|&修复后截图"

            detail_list = []
            for detail in review_record_details:
                detail_split = detail.split('&|&')
                if code_format.startswith(backend_code_format1):
                    detail_list.append(
                        {
                            "code_filepath": detail_split[4],     # 文件路径
                            "code_linenumber": detail_split[5],   # 代码行号
                            "code_snippet": detail_split[6],      # 代码片段
                            "comment_type": detail_split[7],      # 意见类型
                            "checked_by": detail_split[8],        # 检视人员
                            "comment": detail_split[10],          # 检视意见
                            "confirmed_by": detail_split[12],     # 实际确认人员
                            "confirm_result": detail_split[13]    # 确认结果
                        }
                    )
                elif code_format.startswith(backend_code_format2):
                    detail_list.append(
                        {
                            "code_filepath": detail_split[2],     # 文件路径
                            "code_linenumber": detail_split[3],   # 代码行号
                            "code_snippet": detail_split[4],      # 代码片段
                            "comment_type": detail_split[5],      # 意见类型
                            "checked_by": detail_split[6],        # 检视人员
                            "comment": detail_split[8],           # 检视意见
                            "confirmed_by": detail_split[10],     # 实际确认人员
                            "confirm_result": detail_split[11]    # 确认结果
                        }
                    )
                elif code_format == frontend_code_format:
                    detail_list.append(
                        {
                            "code_filepath": detail_split[1],
                            "code_linenumber": detail_split[2],
                            "code_snippet": detail_split[3],
                            "comment_type": detail_split[4],
                            "checked_by": detail_split[5],
                            "comment": detail_split[7],
                            "confirmed_by": detail_split[8],
                            "confirm_result": detail_split[9]
                        }
                    )
                else:
                    print(f"未知的表头格式: {code_format}, 文件URL: {url}")
            return detail_list
        except Exception as e:
            print(f"读取Excel文件时发生错误: {e}")
    else:
        print(f"无法下载文件, HTTP状态码: {response.status_code}, 错误信息: {response.text}")


def get_page_content(page_id):
    """获取页面内容"""
    url = f'{CONFLUENCE_URL}/rest/api/content/{page_id}?expand=body.storage,attachments'
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print("Failed to retrieve page content")
        return None


def parse_records_from_page(project, soup):
    """解析评审记录"""

    def extract_sprint_number(input_text):
        """提取sprint编号"""
        import re
        # 使用正则表达式查找sprint后面的数字，忽略大小写
        match = re.search(r'sprint(\d+)', input_text, re.IGNORECASE)
        if match:
            # 返回统一格式的字符串：Sprint加上匹配到的数字
            return f"Sprint{match.group(1)}"
        else:
            return None

    def parse_record(sub_paragraphs):
        """解析评审记录"""
        title = sprint = review_type = review_date = reviewers = ""
        attachments = []
        for line in sub_paragraphs:
            _first_child = line.contents[0] if line.contents else None
            _is_attachment = _first_child and _first_child.name == 'ac:link'
            if not _is_attachment and '代码评审' in line.text:
                title = line.text
                sprint = extract_sprint_number(title)
                # 如果包含'前端代码评审',设置'前端'; 如果包含'后端代码评审',设置为'后端'; 否则设置为None
                review_type = '前端' if '前端代码评审' in line.text else ('后端' if '后端代码评审' in line.text else None)
            elif '日期' in line.text or '时间' in line.text:
                review_date = line.text.replace('：', ':').replace('日期:', '').replace('时间:', '')
            elif '评审人员' in line.text or '参与人' in line.text:
                reviewers = line.text.replace('：', ':').split(":", 1)[1].strip()
            elif _first_child and _first_child.name == 'ac:link':
                filename = _first_child.find('ri:attachment')['ri:filename']
                if filename.split('.')[-1] == 'xlsx':
                    attachments.append(filename)
                    if not review_type:
                        review_type = '前端' if '前端' in filename else ('后端' if '后端' in filename else None)
        if attachments:
            return {"project": project,
                    "sprint": sprint,
                    "title": title,
                    "review_type": review_type,
                    "review_date": review_date,
                    "reviewers": reviewers,
                    "attachments": attachments
                    }
        else:
            return None

    paragraphs = soup.find_all('p')
    records = []
    flag, i, current_start_line = 0, 0, 0  # flag用于标记解析到的代码评审的行是否是开始行
    while i < len(paragraphs):
        # 判断是否是附件
        first_child = paragraphs[i].contents[0] if paragraphs[i].contents else None
        is_attachment = first_child and first_child.name == 'ac:link'
        # 是否是符合要求的代码评审（是代码评审 and 不是ota升级）
        text = paragraphs[i].text.strip()
        is_code_review = ('代码评审' in text and 'ota升级' not in text)
        if not is_attachment and is_code_review and flag == 0:
            current_start_line = i
            flag = 1 # 解析到开始行，标记为1
        elif not is_attachment and is_code_review and flag == 1:
            record = parse_record(paragraphs[current_start_line: i])
            if record:
                records.append(record)
            flag = 0 # 解析到结束行，标记为0
            i -= 1
        elif i == len(paragraphs) - 1: # 解析到最后一行
            record = parse_record(paragraphs[current_start_line: i+1])
            if record:
                records.append(record)
            flag = 0
        i += 1
    return records


def main():
    # 清空数据表
    print("开始清空数据表: review_records_details ...")
    DbOperator.truncate_table('review_records_details')
    print("完成数据表清理\n")

    for page_id, page_name in PAGE_IDS.items():
        print(f"正在处理 {page_name}...")

        # 获取页面内容
        page_content = get_page_content(page_id)
        if page_content:
            body_content = page_content['body']['storage']['value']
            soup = BeautifulSoup(body_content, 'html.parser')
            records = parse_records_from_page(page_name, soup)

            records_and_details = []
            for record in records:
                for filename in record.get('attachments'):
                    sprint_no = int(record.get('sprint').replace('Sprint', ''))
                    # 解析xlsx文件，sprint编号大于10或页面ID为104896273（数据工具链平台）
                    if filename.split('.')[-1] == 'xlsx' and (sprint_no > 10 or page_id == '104896273'):
                        fileurl = "http://wiki.zvos.zoomlion.com/download/attachments/{}/{}".format(page_id, filename)
                        details = parse_detail_from_xlsx(fileurl)
                        records_and_details.extend(
                            (
                                {**record, 'attachments': '\n'.join(record['attachments']), **detail}
                                for detail in details
                            )
                        )

            # 将解析出来的记录保存到数据库中
            records_df = pd.DataFrame(records_and_details)
            records_df.to_sql('review_records_details', con=DbOperator.get_engine(), if_exists='append', index=False)
        else:
            print("Could not fetch the page content.")


if __name__ == '__main__':
    main()
